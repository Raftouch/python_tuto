import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # cell = sheet['a1'] # column A, row 1
    cell = sheet.cell(1,1) # row 1, column 1
    print(cell.value)
    # print(sheet.max_row)

    for row in range(2, sheet.max_row + 1):
        # print(raw)
        cell = sheet.cell(row, 3)
        # print(cell.value)
        print(f"Row {row}, Value: {cell.value}, Type: {type(cell.value)}")

        if isinstance(cell.value, str):
            try:
                cell.value = float(cell.value)
            except ValueError:
                print(f'Skip non-numeric value in row {row}: {cell.value}')

        if isinstance(cell.value, (int, float)):
            corrected_price = cell.value * 0.9
            # print(corrected_price)
            corrected_price_cell = sheet.cell(row, 4)
            corrected_price_cell.value = corrected_price
        else:
            print(f'Skip non-numeric value in row {row}: {cell.value}')

    values = Reference(sheet, 
                    min_row=2, 
                    max_row=sheet.max_row, 
                    min_col=4, 
                    max_col=4)

    chart = BarChart()
    chart.add_data(values)

    sheet.add_chart(chart, 'e2')

    wb.save(filename)


process_workbook('transactions.xlsx')

